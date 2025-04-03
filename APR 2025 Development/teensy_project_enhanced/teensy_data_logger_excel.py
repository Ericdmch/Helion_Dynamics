#!/usr/bin/env python3
"""
Teensy Data Logger with Excel Integration

This Python application receives data from a Teensy 4.1 via serial connection,
processes it, logs it to CSV/JSON files, and creates Excel spreadsheets with graphs.
It's designed to work on both Mac and Windows.

The data includes:
- BME680 environmental sensor readings (temperature, humidity, pressure, gas)
- Analog photodiode module readings
- MPU6050 accelerometer/gyroscope readings
"""

import serial
import serial.tools.list_ports
import time
import datetime
import csv
import os
import argparse
import platform
import json
import threading
import sys
from pathlib import Path
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.chart import (
    LineChart,
    Reference,
    Series,
)
from openpyxl.chart.axis import DateAxis
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image

# Default settings
DEFAULT_BAUD_RATE = 115200
DEFAULT_LOG_DIR = str(Path.home() / "Desktop/Teensy_Data_Logs")  # Changed to Desktop
DEFAULT_CSV_FILENAME = "sensor_data.csv"
DEFAULT_JSON_FILENAME = "sensor_data.json"
DEFAULT_EXCEL_FILENAME = "sensor_data.xlsx"
DEFAULT_GRAPH_DIR = "graphs"

class TeensyDataLogger:
    def __init__(self, port=None, baud_rate=DEFAULT_BAUD_RATE, log_dir=DEFAULT_LOG_DIR):
        """Initialize the Teensy Data Logger."""
        self.port = port
        self.baud_rate = baud_rate
        self.log_dir = log_dir
        self.graph_dir = os.path.join(log_dir, DEFAULT_GRAPH_DIR)
        self.serial_connection = None
        self.running = False
        self.data_count = 0
        self.data_buffer = []
        self.excel_update_interval = 10  # Update Excel file every 10 data points
        
        # Create log and graph directories if they don't exist
        Path(log_dir).mkdir(parents=True, exist_ok=True)
        Path(self.graph_dir).mkdir(parents=True, exist_ok=True)
        
        # Initialize file paths
        self.csv_path = os.path.join(log_dir, DEFAULT_CSV_FILENAME)
        self.json_path = os.path.join(log_dir, DEFAULT_JSON_FILENAME)
        self.excel_path = os.path.join(log_dir, DEFAULT_EXCEL_FILENAME)
        
        # Check if CSV file exists, if not create it with headers
        csv_exists = os.path.isfile(self.csv_path)
        with open(self.csv_path, 'a', newline='') as csvfile:
            writer = csv.writer(csvfile)
            if not csv_exists:
                writer.writerow([
                    'Timestamp', 'Local Time', 'Device Time (ms)', 'Sender Address', 
                    'RSSI', 'SNR', 'Temperature (°C)', 'Pressure (hPa)', 'Humidity (%)', 
                    'Gas Resistance (kOhms)', 'AccelX (g)', 'AccelY (g)', 'AccelZ (g)', 
                    'GyroX (°/s)', 'GyroY (°/s)', 'GyroZ (°/s)', 'Light Level'
                ])
        
        # Initialize Excel file if it doesn't exist
        if not os.path.isfile(self.excel_path):
            self.create_excel_file()
    
    def find_teensy_port(self):
        """Automatically find the Teensy serial port."""
        system = platform.system()
        ports = list(serial.tools.list_ports.comports())
        
        for port in ports:
            # Look for Teensy in the description
            if "Teensy" in port.description:
                return port.device
            
            # On Mac, Teensy often appears as "usbmodem"
            if system == "Darwin" and "usbmodem" in port.device:
                return port.device
                
            # On Windows, check for specific VID:PID combinations
            if system == "Windows" and hasattr(port, 'vid') and hasattr(port, 'pid'):
                # Teensy 4.1 VID:PID is typically 0x16C0:0x0483
                if port.vid == 0x16C0 and port.pid == 0x0483:
                    return port.device
        
        return None
    
    def connect(self):
        """Connect to the Teensy device."""
        if not self.port:
            self.port = self.find_teensy_port()
            if not self.port:
                raise Exception("Teensy device not found. Please specify port manually.")
        
        try:
            self.serial_connection = serial.Serial(self.port, self.baud_rate, timeout=1)
            print(f"Connected to Teensy on {self.port} at {self.baud_rate} baud")
            return True
        except serial.SerialException as e:
            print(f"Error connecting to Teensy: {e}")
            return False
    
    def disconnect(self):
        """Disconnect from the Teensy device."""
        if self.serial_connection and self.serial_connection.is_open:
            self.serial_connection.close()
            print("Disconnected from Teensy")
    
    def parse_data(self, data_string):
        """Parse the data string received from Teensy."""
        # Expected format: TIME:123456,ADDR:1,RSSI:-80,SNR:10,T:25.50,P:1013.25,H:45.20,G:100.50,AX:0.10,AY:0.20,AZ:0.98,GX:1.50,GY:2.50,GZ:0.50,L:512
        data_dict = {}
        
        # Split by comma and process each key-value pair
        for item in data_string.split(','):
            if ':' in item:
                key, value = item.split(':', 1)
                try:
                    # Try to convert to float if possible
                    data_dict[key] = float(value)
                except ValueError:
                    # Keep as string if not a number
                    data_dict[key] = value
        
        return data_dict
    
    def log_data(self, data_dict):
        """Log the parsed data to CSV, JSON, and buffer for Excel."""
        timestamp = time.time()
        local_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Create a row for CSV
        csv_row = [
            timestamp,
            local_time,
            data_dict.get('TIME', ''),
            data_dict.get('ADDR', ''),
            data_dict.get('RSSI', ''),
            data_dict.get('SNR', ''),
            data_dict.get('T', ''),
            data_dict.get('P', ''),
            data_dict.get('H', ''),
            data_dict.get('G', ''),
            data_dict.get('AX', ''),
            data_dict.get('AY', ''),
            data_dict.get('AZ', ''),
            data_dict.get('GX', ''),
            data_dict.get('GY', ''),
            data_dict.get('GZ', ''),
            data_dict.get('L', '')
        ]
        
        # Log to CSV
        with open(self.csv_path, 'a', newline='') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(csv_row)
        
        # Log to JSON (append to JSON array)
        json_entry = {
            'timestamp': timestamp,
            'local_time': local_time,
            'data': data_dict
        }
        
        # Append to JSON file
        with open(self.json_path, 'a') as jsonfile:
            jsonfile.write(json.dumps(json_entry) + '\n')
        
        # Add to data buffer for Excel
        self.data_buffer.append({
            'Timestamp': timestamp,
            'Local Time': local_time,
            'Device Time (ms)': data_dict.get('TIME', ''),
            'Sender Address': data_dict.get('ADDR', ''),
            'RSSI': data_dict.get('RSSI', ''),
            'SNR': data_dict.get('SNR', ''),
            'Temperature (°C)': data_dict.get('T', ''),
            'Pressure (hPa)': data_dict.get('P', ''),
            'Humidity (%)': data_dict.get('H', ''),
            'Gas Resistance (kOhms)': data_dict.get('G', ''),
            'AccelX (g)': data_dict.get('AX', ''),
            'AccelY (g)': data_dict.get('AY', ''),
            'AccelZ (g)': data_dict.get('AZ', ''),
            'GyroX (°/s)': data_dict.get('GX', ''),
            'GyroY (°/s)': data_dict.get('GY', ''),
            'GyroZ (°/s)': data_dict.get('GZ', ''),
            'Light Level': data_dict.get('L', '')
        })
        
        self.data_count += 1
        
        # Update Excel file periodically
        if self.data_count % self.excel_update_interval == 0:
            self.update_excel_file()
            self.generate_graphs()
        
        return True
    
    def display_data(self, data_dict):
        """Display the parsed data in a readable format."""
        print("\n--- Received Data ---")
        print(f"Time: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"Device Time: {data_dict.get('TIME', 'N/A')} ms")
        print(f"Sender: {data_dict.get('ADDR', 'N/A')}")
        print(f"Signal: RSSI={data_dict.get('RSSI', 'N/A')} dBm, SNR={data_dict.get('SNR', 'N/A')} dB")
        print("\nEnvironmental Data (BME680):")
        print(f"  Temperature: {data_dict.get('T', 'N/A')} °C")
        print(f"  Pressure: {data_dict.get('P', 'N/A')} hPa")
        print(f"  Humidity: {data_dict.get('H', 'N/A')} %")
        print(f"  Gas Resistance: {data_dict.get('G', 'N/A')} kOhms")
        print("\nMotion Data (MPU6050):")
        print(f"  Acceleration: X={data_dict.get('AX', 'N/A')}, Y={data_dict.get('AY', 'N/A')}, Z={data_dict.get('AZ', 'N/A')} g")
        print(f"  Gyroscope: X={data_dict.get('GX', 'N/A')}, Y={data_dict.get('GY', 'N/A')}, Z={data_dict.get('GZ', 'N/A')} °/s")
        print("\nLight Data (Photodiode):")
        print(f"  Light Level: {data_dict.get('L', 'N/A')}")
        print(f"Total records: {self.data_count}")
        print("---------------------")
    
    def create_excel_file(self):
        """Create a new Excel file with proper formatting and headers."""
        wb = Workbook()
        
        # Create Data Sheet
        ws_data = wb.active
        ws_data.title = "Sensor Data"
        
        # Add headers with formatting
        headers = [
            'Timestamp', 'Local Time', 'Device Time (ms)', 'Sender Address', 
            'RSSI', 'SNR', 'Temperature (°C)', 'Pressure (hPa)', 'Humidity (%)', 
            'Gas Resistance (kOhms)', 'AccelX (g)', 'AccelY (g)', 'AccelZ (g)', 
            'GyroX (°/s)', 'GyroY (°/s)', 'GyroZ (°/s)', 'Light Level'
        ]
        
        # Apply header formatting
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            # Set column width
            ws_data.column_dimensions[get_column_letter(col_num)].width = 15
        
        # Create Dashboard Sheet
        ws_dashboard = wb.create_sheet(title="Dashboard")
        ws_dashboard.cell(row=1, column=1, value="Teensy Sensor Network Dashboard")
        ws_dashboard.cell(row=1, column=1).font = Font(size=16, bold=True)
        ws_dashboard.merge_cells('A1:H1')
        
        # Create sheets for different graph categories
        wb.create_sheet(title="Environmental Graphs")
        wb.create_sheet(title="Motion Graphs")
        wb.create_sheet(title="Light Graphs")
        
        # Save the workbook
        wb.save(self.excel_path)
        print(f"Created Excel file: {self.excel_path}")
    
    def update_excel_file(self):
        """Update the Excel file with new data."""
        try:
            # Load existing data from CSV to ensure we have everything
            df = pd.read_csv(self.csv_path)
            
            # Create a new workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Sensor Data"
            
            # Add headers
            headers = list(df.columns)
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                # Apply header formatting
                cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                # Set column width
                ws.column_dimensions[get_column_letter(col_num)].width = 15
            
            # Add data
            for row_idx, row in df.iterrows():
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx+2, column=col_idx, value=value)
            
            # Create Dashboard Sheet
            ws_dashboard = wb.create_sheet(title="Dashboard")
            ws_dashboard.cell(row=1, column=1, value="Teensy Sensor Network Dashboard")
            ws_dashboard.cell(row=1, column=1).font = Font(size=16, bold=True)
            ws_dashboard.merge_cells('A1:H1')
            
            # Add summary statistics
            ws_dashboard.cell(row=3, column=1, value="Summary Statistics")
            ws_dashboard.cell(row=3, column=1).font = Font(bold=True)
            
            # Add latest values
            if not df.empty:
                latest = df.iloc[-1]
                ws_dashboard.cell(row=5, column=1, value="Latest Readings:")
                ws_dashboard.cell(row=5, column=1).font = Font(bold=True)
                
                ws_dashboard.cell(row=6, column=1, value="Time:")
                ws_dashboard.cell(row=6, column=2, value=latest['Local Time'])
                
                ws_dashboard.cell(row=7, column=1, value="Temperature:")
                ws_dashboard.cell(row=7, column=2, value=f"{latest['Temperature (°C)']} °C")
                
                ws_dashboard.cell(row=8, column=1, value="Humidity:")
                ws_dashboard.cell(row=8, column=2, value=f"{latest['Humidity (%)']} %")
                
                ws_dashboard.cell(row=9, column=1, value="Pressure:")
                ws_dashboard.cell(row=9, column=2, value=f"{latest['Pressure (hPa)']} hPa")
                
                ws_dashboard.cell(row=10, column=1, value="Light Level:")
                ws_dashboard.cell(row=10, column=2, value=f"{latest['Light Level']}")
                
                # Add min/max/avg statistics
                ws_dashboard.cell(row=5, column=4, value="Statistics:")
                ws_dashboard.cell(row=5, column=4).font = Font(bold=True)
                
                ws_dashboard.cell(row=6, column=4, value="Temperature Range:")
                ws_dashboard.cell(row=6, column=5, value=f"{df['Temperature (°C)'].min():.2f} - {df['Temperature (°C)'].max():.2f} °C")
                
                ws_dashboard.cell(row=7, column=4, value="Avg Temperature:")
                ws_dashboard.cell(row=7, column=5, value=f"{df['Temperature (°C)'].mean():.2f} °C")
                
                ws_dashboard.cell(row=8, column=4, value="Avg Humidity:")
                ws_dashboard.cell(row=8, column=5, value=f"{df['Humidity (%)'].mean():.2f} %")
                
                ws_dashboard.cell(row=9, column=4, value="Avg Pressure:")
                ws_dashboard.cell(row=9, column=5, value=f"{df['Pressure (hPa)'].mean():.2f} hPa")
                
                ws_dashboard.cell(row=10, column=4, value="Records Count:")
                ws_dashboard.cell(row=10, column=5, value=f"{len(df)}")
            
            # Create sheets for different graph categories
            wb.create_sheet(title="Environmental Graphs")
            wb.create_sheet(title="Motion Graphs")
            wb.create_sheet(title="Light Graphs")
            
            # Add graph images if they exist
            self.add_graph_images_to_excel(wb)
            
            # Save the workbook
            wb.save(self.excel_path)
            print(f"Updated Excel file with {len(df)} records")
            
        except Exception as e:
            print(f"Error updating Excel file: {e}")
    
    def generate_graphs(self):
        """Generate graphs from the collected data."""
        try:
            # Load data from CSV
            df = pd.read_csv(self.csv_path)
            
            if len(df) < 2:
                print("Not enough data to generate graphs")
                return
            
            # Convert timestamp to datetime for better x-axis
            df['Datetime'] = pd.to_datetime(df['Local Time'])
            
            # Set up the plotting style
            plt.style.use('ggplot')
            
            # 1. Environmental Graphs
            self.generate_environmental_graphs(df)
            
            # 2. Motion Graphs
            self.generate_motion_graphs(df)
            
            # 3. Light Level Graph
            self.generate_light_graph(df)
            
            print(f"Generated graphs in {self.graph_dir}")
            
        except Exception as e:
            print(f"Error generating graphs: {e}")
    
    def generate_environmental_graphs(self, df):
        """Generate graphs for environmental data (BME680)."""
        # Temperature graph
        plt.figure(figsize=(10, 6))
        plt.plot(df['Datetime'], df['Temperature (°C)'], 'r-', linewidth=2)
        plt.title('Temperature Over Time')
        plt.xlabel('Time')
        plt.ylabel('Temperature (°C)')
        plt.grid(True)
        plt.tight_layout()
        plt.savefig(os.path.join(self.graph_dir, 'temperature.png'))
        plt.close()
        
        # Humidity graph
        plt.figure(figsize=(10, 6))
        plt.plot(df['Datetime'], df['Humidity (%)'], 'b-', linewidth=2)
        plt.title('Humidity Over Time')
        plt.xlabel('Time')
        plt.ylabel('Humidity (%)')
        plt.grid(True)
        plt.tight_layout()
        plt.savefig(os.path.join(self.graph_dir, 'humidity.png'))
        plt.close()
        
        # Pressure graph
        plt.figure(figsize=(10, 6))
        plt.plot(df['Datetime'], df['Pressure (hPa)'], 'g-', linewidth=2)
        plt.title('Atmospheric Pressure Over Time')
        plt.xlabel('Time')
        plt.ylabel('Pressure (hPa)')
        plt.grid(True)
        plt.tight_layout()
        plt.savefig(os.path.join(self.graph_dir, 'pressure.png'))
        plt.close()
        
        # Gas Resistance graph
        plt.figure(figsize=(10, 6))
        plt.plot(df['Datetime'], df['Gas Resistance (kOhms)'], 'y-', linewidth=2)
        plt.title('Gas Resistance Over Time')
        plt.xlabel('Time')
        plt.ylabel('Gas Resistance (kOhms)')
        plt.grid(True)
        plt.tight_layout()
        plt.savefig(os.path.join(self.graph_dir, 'gas.png'))
        plt.close()
        
        # Combined environmental graph
        plt.figure(figsize=(12, 8))
        
        # Create subplots
        ax1 = plt.subplot(2, 2, 1)
        ax1.plot(df['Datetime'], df['Temperature (°C)'], 'r-')
        ax1.set_title('Temperature (°C)')
        ax1.grid(True)
        
        ax2 = plt.subplot(2, 2, 2)
        ax2.plot(df['Datetime'], df['Humidity (%)'], 'b-')
        ax2.set_title('Humidity (%)')
        ax2.grid(True)
        
        ax3 = plt.subplot(2, 2, 3)
        ax3.plot(df['Datetime'], df['Pressure (hPa)'], 'g-')
        ax3.set_title('Pressure (hPa)')
        ax3.grid(True)
        
        ax4 = plt.subplot(2, 2, 4)
        ax4.plot(df['Datetime'], df['Gas Resistance (kOhms)'], 'y-')
        ax4.set_title('Gas Resistance (kOhms)')
        ax4.grid(True)
        
        plt.tight_layout()
        plt.savefig(os.path.join(self.graph_dir, 'environmental_combined.png'))
        plt.close()
    
    def generate_motion_graphs(self, df):
        """Generate graphs for motion data (MPU6050)."""
        # Acceleration graph
        plt.figure(figsize=(10, 6))
        plt.plot(df['Datetime'], df['AccelX (g)'], 'r-', label='X')
        plt.plot(df['Datetime'], df['AccelY (g)'], 'g-', label='Y')
        plt.plot(df['Datetime'], df['AccelZ (g)'], 'b-', label='Z')
        plt.title('Acceleration Over Time')
        plt.xlabel('Time')
        plt.ylabel('Acceleration (g)')
        plt.legend()
        plt.grid(True)
        plt.tight_layout()
        plt.savefig(os.path.join(self.graph_dir, 'acceleration.png'))
        plt.close()
        
        # Gyroscope graph
        plt.figure(figsize=(10, 6))
        plt.plot(df['Datetime'], df['GyroX (°/s)'], 'r-', label='X')
        plt.plot(df['Datetime'], df['GyroY (°/s)'], 'g-', label='Y')
        plt.plot(df['Datetime'], df['GyroZ (°/s)'], 'b-', label='Z')
        plt.title('Gyroscope Readings Over Time')
        plt.xlabel('Time')
        plt.ylabel('Angular Velocity (°/s)')
        plt.legend()
        plt.grid(True)
        plt.tight_layout()
        plt.savefig(os.path.join(self.graph_dir, 'gyroscope.png'))
        plt.close()
        
        # Combined motion graph (3D scatter)
        fig = plt.figure(figsize=(10, 8))
        ax = fig.add_subplot(111, projection='3d')
        
        # Get the last 100 points or all if less than 100
        points_to_plot = min(100, len(df))
        recent_df = df.tail(points_to_plot)
        
        ax.scatter(recent_df['AccelX (g)'], recent_df['AccelY (g)'], recent_df['AccelZ (g)'], 
                  c=range(len(recent_df)), cmap='viridis', marker='o')
        
        ax.set_xlabel('X Acceleration (g)')
        ax.set_ylabel('Y Acceleration (g)')
        ax.set_zlabel('Z Acceleration (g)')
        ax.set_title('3D Acceleration Plot (Recent Data)')
        
        plt.tight_layout()
        plt.savefig(os.path.join(self.graph_dir, 'acceleration_3d.png'))
        plt.close()
    
    def generate_light_graph(self, df):
        """Generate graph for light level data (photodiode)."""
        plt.figure(figsize=(10, 6))
        plt.plot(df['Datetime'], df['Light Level'], 'y-', linewidth=2)
        plt.title('Light Level Over Time')
        plt.xlabel('Time')
        plt.ylabel('Light Level')
        plt.grid(True)
        plt.tight_layout()
        plt.savefig(os.path.join(self.graph_dir, 'light_level.png'))
        plt.close()
    
    def add_graph_images_to_excel(self, workbook):
        """Add generated graph images to the Excel workbook."""
        try:
            # Environmental Graphs sheet
            ws_env = workbook["Environmental Graphs"]
            ws_env.column_dimensions['A'].width = 5
            ws_env.column_dimensions['B'].width = 30
            
            # Add title
            ws_env.cell(row=1, column=2, value="Environmental Sensor Graphs")
            ws_env.cell(row=1, column=2).font = Font(size=14, bold=True)
            
            # Add temperature graph
            temp_img_path = os.path.join(self.graph_dir, 'temperature.png')
            if os.path.exists(temp_img_path):
                img = Image(temp_img_path)
                img.width = 600
                img.height = 350
                ws_env.add_image(img, 'B3')
                ws_env.cell(row=3, column=1, value="1")
                ws_env.cell(row=3, column=2, value="Temperature Over Time")
                ws_env.cell(row=3, column=2).font = Font(bold=True)
            
            # Add humidity graph
            humidity_img_path = os.path.join(self.graph_dir, 'humidity.png')
            if os.path.exists(humidity_img_path):
                img = Image(humidity_img_path)
                img.width = 600
                img.height = 350
                ws_env.add_image(img, 'B20')
                ws_env.cell(row=20, column=1, value="2")
                ws_env.cell(row=20, column=2, value="Humidity Over Time")
                ws_env.cell(row=20, column=2).font = Font(bold=True)
            
            # Add pressure graph
            pressure_img_path = os.path.join(self.graph_dir, 'pressure.png')
            if os.path.exists(pressure_img_path):
                img = Image(pressure_img_path)
                img.width = 600
                img.height = 350
                ws_env.add_image(img, 'B37')
                ws_env.cell(row=37, column=1, value="3")
                ws_env.cell(row=37, column=2, value="Pressure Over Time")
                ws_env.cell(row=37, column=2).font = Font(bold=True)
            
            # Motion Graphs sheet
            ws_motion = workbook["Motion Graphs"]
            ws_motion.column_dimensions['A'].width = 5
            ws_motion.column_dimensions['B'].width = 30
            
            # Add title
            ws_motion.cell(row=1, column=2, value="Motion Sensor Graphs")
            ws_motion.cell(row=1, column=2).font = Font(size=14, bold=True)
            
            # Add acceleration graph
            accel_img_path = os.path.join(self.graph_dir, 'acceleration.png')
            if os.path.exists(accel_img_path):
                img = Image(accel_img_path)
                img.width = 600
                img.height = 350
                ws_motion.add_image(img, 'B3')
                ws_motion.cell(row=3, column=1, value="1")
                ws_motion.cell(row=3, column=2, value="Acceleration Over Time")
                ws_motion.cell(row=3, column=2).font = Font(bold=True)
            
            # Add gyroscope graph
            gyro_img_path = os.path.join(self.graph_dir, 'gyroscope.png')
            if os.path.exists(gyro_img_path):
                img = Image(gyro_img_path)
                img.width = 600
                img.height = 350
                ws_motion.add_image(img, 'B20')
                ws_motion.cell(row=20, column=1, value="2")
                ws_motion.cell(row=20, column=2, value="Gyroscope Readings Over Time")
                ws_motion.cell(row=20, column=2).font = Font(bold=True)
            
            # Add 3D acceleration graph
            accel_3d_img_path = os.path.join(self.graph_dir, 'acceleration_3d.png')
            if os.path.exists(accel_3d_img_path):
                img = Image(accel_3d_img_path)
                img.width = 600
                img.height = 350
                ws_motion.add_image(img, 'B37')
                ws_motion.cell(row=37, column=1, value="3")
                ws_motion.cell(row=37, column=2, value="3D Acceleration Plot")
                ws_motion.cell(row=37, column=2).font = Font(bold=True)
            
            # Light Graphs sheet
            ws_light = workbook["Light Graphs"]
            ws_light.column_dimensions['A'].width = 5
            ws_light.column_dimensions['B'].width = 30
            
            # Add title
            ws_light.cell(row=1, column=2, value="Light Sensor Graphs")
            ws_light.cell(row=1, column=2).font = Font(size=14, bold=True)
            
            # Add light level graph
            light_img_path = os.path.join(self.graph_dir, 'light_level.png')
            if os.path.exists(light_img_path):
                img = Image(light_img_path)
                img.width = 600
                img.height = 350
                ws_light.add_image(img, 'B3')
                ws_light.cell(row=3, column=1, value="1")
                ws_light.cell(row=3, column=2, value="Light Level Over Time")
                ws_light.cell(row=3, column=2).font = Font(bold=True)
            
            # Add combined environmental graph to dashboard
            ws_dashboard = workbook["Dashboard"]
            combined_env_img_path = os.path.join(self.graph_dir, 'environmental_combined.png')
            if os.path.exists(combined_env_img_path):
                img = Image(combined_env_img_path)
                img.width = 600
                img.height = 400
                ws_dashboard.add_image(img, 'A12')
                ws_dashboard.cell(row=12, column=1, value="Environmental Data Overview:")
                ws_dashboard.cell(row=12, column=1).font = Font(bold=True)
            
        except Exception as e:
            print(f"Error adding graph images to Excel: {e}")
    
    def start_logging(self):
        """Start the logging process."""
        if not self.serial_connection or not self.serial_connection.is_open:
            if not self.connect():
                return False
        
        self.running = True
        print("Logging started. Press Ctrl+C to stop.")
        print(f"Data will be logged to CSV, JSON, and Excel formats in the '{self.log_dir}' directory.")
        print(f"Graphs will be generated in the '{self.graph_dir}' directory.")
        
        try:
            while self.running:
                if self.serial_connection.in_waiting:
                    line = self.serial_connection.readline().decode('utf-8').strip()
                    if line:
                        try:
                            data_dict = self.parse_data(line)
                            if data_dict:
                                self.log_data(data_dict)
                                self.display_data(data_dict)
                        except Exception as e:
                            print(f"Error processing data: {e}")
                            print(f"Raw data: {line}")
                time.sleep(0.01)  # Small delay to prevent CPU hogging
        except KeyboardInterrupt:
            print("\nLogging stopped by user")
        except Exception as e:
            print(f"Error during logging: {e}")
        finally:
            self.running = False
            # Make sure to update Excel with any remaining data
            if self.data_buffer:
                self.update_excel_file()
                self.generate_graphs()
            self.disconnect()
        
        return True

def list_serial_ports():
    """List all available serial ports."""
    ports = list(serial.tools.list_ports.comports())
    if not ports:
        print("No serial ports found")
        return
    
    print("\nAvailable serial ports:")
    for i, port in enumerate(ports):
        print(f"{i+1}. {port.device} - {port.description}")

def main():
    """Main function to parse arguments and start the logger."""
    parser = argparse.ArgumentParser(description='Teensy Data Logger with Excel Integration')
    parser.add_argument('-p', '--port', help='Serial port to connect to')
    parser.add_argument('-b', '--baud', type=int, default=DEFAULT_BAUD_RATE, help='Baud rate')
    parser.add_argument('-d', '--directory', default=DEFAULT_LOG_DIR, help='Log directory')
    parser.add_argument('-l', '--list', action='store_true', help='List available serial ports')
    
    args = parser.parse_args()
    
    if args.list:
        list_serial_ports()
        return
    
    logger = TeensyDataLogger(port=args.port, baud_rate=args.baud, log_dir=args.directory)
    
    try:
        logger.start_logging()
    except KeyboardInterrupt:
        print("\nProgram terminated by user")
    except Exception as e:
        print(f"Error: {e}")
    finally:
        logger.disconnect()

if __name__ == "__main__":
    main()
